"""
Excel styling utilities.
"""

from functools import lru_cache
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


@lru_cache(maxsize=64)
def create_fill(color: str) -> PatternFill:
    """Create a solid fill pattern (cached for performance)."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def create_thin_border() -> Border:
    """Create a thin border for cells."""
    return Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )


def create_header_style(header_color: str) -> dict:
    """Create header cell style."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": create_fill(header_color),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": create_thin_border(),
    }


def apply_header_style(cell, header_color: str):
    """Apply header style to a cell."""
    style = create_header_style(header_color)
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]


def get_level_fill(
    level: str | int, level_colors: dict[str, str]
) -> PatternFill | None:
    """Get fill color for a skill level."""
    str_level = str(level)
    if str_level in level_colors:
        return create_fill(level_colors[str_level])
    return None


def get_category_fill(
    category_id: str, category_colors: dict[str, str]
) -> PatternFill | None:
    """Get fill color for a category."""
    if category_id in category_colors:
        return create_fill(category_colors[category_id])
    return None
