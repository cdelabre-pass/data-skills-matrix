"""
Excel generation for skills matrix.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .models import Config, Category, Resource
from .styles import (
    apply_header_style,
    get_level_fill,
    get_category_fill,
    create_thin_border,
    create_fill,
)


def format_resources_for_excel(resources: list[Resource]) -> str:
    """Format a list of Resource objects as a string for Excel."""
    if not resources:
        return ""
    return "\n".join(f"• {r.title}: {r.url}" for r in resources)


class SheetBuilder:
    """Helper class to build Excel sheets with common operations."""

    def __init__(self, ws: Worksheet, config: Config):
        self.ws = ws
        self.config = config
        self.header_color = config.colors.header

    def add_title(self, text: str, row: int = 1, merge_to_col: int = 1):
        """Add a title row with header styling."""
        cell = self.ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = create_fill(self.header_color)
        if merge_to_col > 1:
            self.ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=merge_to_col
            )

    def add_section_title(self, text: str, row: int):
        """Add a section title (bold, no fill)."""
        cell = self.ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)

    def add_headers(self, headers: list[str], row: int = 1):
        """Add header row with styling."""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            apply_header_style(cell, self.header_color)

    def set_column_widths(self, widths: dict[str, int]):
        """Set column widths from a dict mapping column letters to widths."""
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

    def set_column_widths_range(self, start_col: int, end_col: int, width: int):
        """Set the same width for a range of columns."""
        for col in range(start_col, end_col + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = width


def create_levels_sheet(wb: Workbook, config: Config):
    """Create the skill levels reference sheet."""
    ws = wb.create_sheet("Niveaux", 0)
    builder = SheetBuilder(ws, config)
    level_colors = config.colors.levels

    # Title
    builder.add_title("Échelle de Compétences", merge_to_col=3)

    # Headers
    builder.add_headers(["Niveau", "Description", "Exemple"], row=3)

    # Standard and bonus levels
    all_levels = config.skill_levels.get("standard", []) + config.skill_levels.get(
        "bonus", []
    )

    # Add NC level
    levels_data = [(str(sl.level), sl.name, sl.description) for sl in all_levels]
    levels_data.append(
        ("NC", "Non Concerné", "Compétence non attendue pour ce rôle/niveau")
    )

    for row_idx, (level, desc, example) in enumerate(levels_data, 4):
        ws.cell(row=row_idx, column=1, value=level)
        ws.cell(row=row_idx, column=2, value=desc)
        ws.cell(row=row_idx, column=3, value=example)

        # Level color
        fill = get_level_fill(level, level_colors)
        if fill:
            ws.cell(row=row_idx, column=1).fill = fill

    # Column widths
    builder.set_column_widths({"A": 15, "B": 30, "C": 50})

    # Career levels section
    start_row = 4 + len(levels_data) + 2
    builder.add_section_title("Niveaux attendus par Rôle", start_row)
    builder.add_headers(["Rôle", "Expérience", "Niveaux attendus"], row=start_row + 2)

    for row_idx, cl in enumerate(config.career_levels, start_row + 3):
        ws.cell(row=row_idx, column=1, value=cl.name)
        ws.cell(row=row_idx, column=2, value=cl.experience)
        core_range = f"{cl.core_expected[0]}-{cl.core_expected[1]}"
        sec_range = f"{cl.secondary_expected[0]}-{cl.secondary_expected[1]}"
        ws.cell(
            row=row_idx,
            column=3,
            value=f"Compétences core: Niveau {core_range} | Secondaires: Niveau {sec_range}",
        )

    # Core vs Secondary legend
    legend_row = start_row + 3 + len(config.career_levels) + 2
    builder.add_section_title("Core vs Secondary Skills", legend_row)

    core_color = config.colors.core_secondary["core"]
    secondary_color = config.colors.core_secondary["secondary"]

    ws.cell(row=legend_row + 2, column=1, value="C")
    ws.cell(row=legend_row + 2, column=1).fill = create_fill(core_color)
    ws.cell(row=legend_row + 2, column=1).font = Font(bold=True)
    ws.cell(
        row=legend_row + 2,
        column=2,
        value="Core Skill - Compétence fondamentale pour le rôle",
    )

    ws.cell(row=legend_row + 3, column=1, value="S")
    ws.cell(row=legend_row + 3, column=1).fill = create_fill(secondary_color)
    ws.cell(
        row=legend_row + 3,
        column=2,
        value="Secondary Skill - Compétence complémentaire",
    )


def create_matrix_sheet(wb: Workbook, config: Config, categories: list[Category]):
    """Create the main skills matrix sheet."""
    ws = wb.create_sheet("Matrice Compétences", 1)
    builder = SheetBuilder(ws, config)
    level_colors = config.colors.levels
    category_colors = config.colors.categories
    core_color = config.colors.core_secondary["core"]
    secondary_color = config.colors.core_secondary["secondary"]

    # Headers
    headers = ["Catégorie", "Compétence", "Description"]
    for role in config.roles:
        headers.extend(
            [
                f"{role.name} C/S",
                f"{role.name} Jr",
                f"{role.name} Cf",
                f"{role.name} Sr",
                f"{role.name} Ex",
            ]
        )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell, builder.header_color)
        if col_idx > 3:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True, text_rotation=90
            )

    # Data
    row_idx = 2
    col = len(headers) + 1  # Track last column for width setting
    border = create_thin_border()

    for category in categories:
        for skill in category.skills:
            # Category
            cell = ws.cell(row=row_idx, column=1, value=category.name)
            fill = get_category_fill(category.id, category_colors)
            if fill:
                cell.fill = fill

            # Skill name and description
            ws.cell(row=row_idx, column=2, value=skill.name)
            ws.cell(row=row_idx, column=3, value=skill.description)

            # Levels per role
            col = 4
            for role in config.roles:
                # Core/Secondary
                is_core = role.id in skill.core_roles
                cs_cell = ws.cell(
                    row=row_idx, column=col, value="C" if is_core else "S"
                )
                cs_cell.alignment = Alignment(horizontal="center")
                cs_cell.fill = create_fill(core_color if is_core else secondary_color)
                if is_core:
                    cs_cell.font = Font(bold=True)
                col += 1

                # Levels: Junior, Confirmed, Senior, Expert
                levels = skill.levels.get(role.id, ["NC", "NC", "NC", "NC"])
                for level in levels:
                    cell = ws.cell(row=row_idx, column=col, value=str(level))
                    cell.alignment = Alignment(horizontal="center")
                    str_level = str(level)
                    fill = get_level_fill(str_level, level_colors)
                    if fill:
                        cell.fill = fill
                        if str_level in ["3", "4"]:
                            cell.font = Font(color="FFFFFF", bold=True)
                    col += 1

            # Borders for entire row
            for c in range(1, col):
                ws.cell(row=row_idx, column=c).border = border

            row_idx += 1

    # Column widths
    builder.set_column_widths({"A": 25, "B": 35, "C": 60})
    builder.set_column_widths_range(4, col - 1, 8)

    # Freeze panes
    ws.freeze_panes = "D2"


def create_level_descriptions_sheet(
    wb: Workbook, config: Config, categories: list[Category]
):
    """Create the detailed level descriptions sheet."""
    ws = wb.create_sheet("Descriptions par Niveau", 2)
    builder = SheetBuilder(ws, config)
    level_colors = config.colors.levels
    category_colors = config.colors.categories

    # Headers
    headers = [
        "Catégorie",
        "Compétence",
        "Niveau 0",
        "Niveau 1",
        "Niveau 2",
        "Niveau 3",
        "Niveau 4",
        "Ressources Internes",
    ]
    builder.add_headers(headers)

    # Data
    row_idx = 2
    border = create_thin_border()

    for category in categories:
        for skill in category.skills:
            # Category with color
            cell = ws.cell(row=row_idx, column=1, value=category.name)
            fill = get_category_fill(category.id, category_colors)
            if fill:
                cell.fill = fill

            # Skill name
            ws.cell(row=row_idx, column=2, value=skill.name)

            # Level descriptions
            for level in range(5):
                desc = skill.level_descriptions.get(level, "")
                cell = ws.cell(row=row_idx, column=3 + level, value=desc)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                fill = get_level_fill(str(level), level_colors)
                if fill:
                    cell.fill = fill
                    if level >= 3:
                        cell.font = Font(color="FFFFFF")

            # Resources
            ws.cell(
                row=row_idx, column=8, value=format_resources_for_excel(skill.resources)
            )

            # Borders
            for c in range(1, 9):
                ws.cell(row=row_idx, column=c).border = border

            row_idx += 1

    # Column widths
    builder.set_column_widths({"A": 25, "B": 30, "H": 50})
    builder.set_column_widths_range(3, 7, 45)

    # Row heights
    for row in range(2, row_idx):
        ws.row_dimensions[row].height = 60

    ws.freeze_panes = "C2"


def create_assessment_sheet(wb: Workbook, config: Config, categories: list[Category]):
    """Create the self-assessment sheet."""
    ws = wb.create_sheet("Mon Évaluation", 3)
    builder = SheetBuilder(ws, config)
    category_colors = config.colors.categories

    # Title and instructions
    builder.add_title("Auto-évaluation des Compétences", merge_to_col=8)

    ws[
        "A2"
    ] = "Instructions: Évaluez votre niveau (0-4) et fournissez des preuves concrètes (liens projets, PRs, dashboards)"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:H2")

    # Role selection
    ws["A3"] = "Mon Rôle:"
    ws["B3"] = ""
    role_names = ",".join(r.name for r in config.roles)
    dv_role = DataValidation(type="list", formula1=f'"{role_names}"')
    ws.add_data_validation(dv_role)
    dv_role.add("B3")

    ws["C3"] = "Mon Niveau:"
    ws["D3"] = ""
    level_names = ",".join(cl.name for cl in config.career_levels)
    dv_level = DataValidation(type="list", formula1=f'"{level_names}"')
    ws.add_data_validation(dv_level)
    dv_level.add("D3")

    # Headers
    headers = [
        "Catégorie",
        "Compétence",
        "Core/Sec",
        "Niveau Attendu",
        "Mon Niveau",
        "Écart",
        "Preuves/Exemples",
        "Commentaires Manager",
    ]
    builder.add_headers(headers, row=5)

    # Skills
    row_idx = 6
    for category in categories:
        for skill in category.skills:
            ws.cell(row=row_idx, column=1, value=category.name)
            ws.cell(row=row_idx, column=2, value=skill.name)
            ws.cell(row=row_idx, column=3, value="")  # Core/Sec - filled based on role
            ws.cell(
                row=row_idx, column=4, value=""
            )  # Expected level - filled based on role
            ws.cell(row=row_idx, column=5, value="")  # Self-assessment
            ws.cell(
                row=row_idx,
                column=6,
                value=f'=IF(E{row_idx}="","",E{row_idx}-D{row_idx})',
            )  # Gap
            ws.cell(row=row_idx, column=7, value="")  # Proof
            ws.cell(row=row_idx, column=8, value="")  # Comments

            # Category color
            fill = get_category_fill(category.id, category_colors)
            if fill:
                ws.cell(row=row_idx, column=1).fill = fill

            row_idx += 1

    # Data validation for My Level (only if there are skills)
    total_skills = sum(len(cat.skills) for cat in categories)
    if total_skills > 0:
        dv = DataValidation(type="list", formula1='"0,1,2,3,4,NC"', allow_blank=True)
        dv.error = "Veuillez entrer un niveau valide (0-4 ou NC)"
        ws.add_data_validation(dv)
        dv.add(f"E6:E{5 + total_skills}")

    # Column widths
    builder.set_column_widths(
        {"A": 25, "B": 35, "C": 10, "D": 12, "E": 12, "F": 8, "G": 50, "H": 40}
    )

    ws.freeze_panes = "A6"


def create_radar_data_sheet(wb: Workbook, config: Config, categories: list[Category]):
    """Create the radar chart data sheet for role profiles."""
    ws = wb.create_sheet("Profils par Rôle (Radar)", 4)
    builder = SheetBuilder(ws, config)

    # Title
    builder.add_title(
        "Profils de Compétences par Rôle (Expert - 10+ ans)", merge_to_col=7
    )

    # Select max 2 skills per category for readability
    selected_skills = []
    for category in categories:
        for skill in category.skills[:2]:
            selected_skills.append((category, skill))

    # Headers
    headers = ["Compétence"] + [role.name for role in config.roles]
    builder.add_headers(headers, row=3)

    # Data (Expert level for each role)
    for row_idx, (category, skill) in enumerate(selected_skills, 4):
        ws.cell(row=row_idx, column=1, value=skill.name)
        for col_idx, role in enumerate(config.roles, 2):
            levels = skill.levels.get(role.id, ["NC", "NC", "NC", "NC"])
            expert_level = levels[3] if len(levels) > 3 and levels[3] != "NC" else 0
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=expert_level if isinstance(expert_level, int) else 0,
            )

    # Create radar charts for each role
    for chart_idx, role in enumerate(config.roles):
        chart = RadarChart()
        chart.type = "filled"
        chart.title = f"Profil {role.name} (Expert)"
        chart.style = 10

        # Chart size (in cm)
        chart.width = 18
        chart.height = 14

        # Data
        labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(selected_skills))
        data = Reference(
            ws, min_col=2 + chart_idx, min_row=3, max_row=3 + len(selected_skills)
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        # Position chart (2 charts per row with larger spacing)
        row_pos = 3 + len(selected_skills) + 2 + (chart_idx // 2) * 22
        col_pos = get_column_letter(1 + (chart_idx % 2) * 12)
        ws.add_chart(chart, f"{col_pos}{row_pos}")

    # Column widths
    builder.set_column_widths({"A": 35})
    builder.set_column_widths_range(2, 7, 15)


def create_development_plan_sheet(wb: Workbook, config: Config):
    """Create the development plan sheet."""
    ws = wb.create_sheet("Plan Développement", 5)
    builder = SheetBuilder(ws, config)

    # Title
    builder.add_title("Plan de Développement Annuel", merge_to_col=8)

    ws[
        "A2"
    ] = "Choisissez un maximum de 4 compétences à développer sur l'année (privilégier les compétences Core)"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:H2")

    # Headers
    headers = [
        "Compétence à Développer",
        "Niveau Actuel",
        "Niveau Cible",
        "Formation / Actions",
        "Résultat Attendu",
        "Mesure de Succès",
        "Trimestre",
        "Statut",
    ]
    builder.add_headers(headers, row=4)

    # Empty rows for 4 skills
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_idx in range(5, 9):
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col, value="")
            cell.border = border

    # Quarter validation
    dv_quarter = DataValidation(type="list", formula1='"Q1,Q2,Q3,Q4"', allow_blank=True)
    ws.add_data_validation(dv_quarter)
    dv_quarter.add("G5:G8")

    # Status validation
    dv_status = DataValidation(
        type="list", formula1='"Non démarré,En cours,Complété"', allow_blank=True
    )
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H8")

    # Column widths
    builder.set_column_widths(
        {"A": 35, "B": 12, "C": 12, "D": 45, "E": 35, "F": 30, "G": 12, "H": 15}
    )


def create_history_sheet(wb: Workbook, config: Config, categories: list[Category]):
    """Create the evaluation history sheet."""
    ws = wb.create_sheet("Historique Évaluations", 6)
    builder = SheetBuilder(ws, config)
    category_colors = config.colors.categories

    # Title
    builder.add_title("Historique des Évaluations", merge_to_col=14)

    # Headers
    headers = [
        "Catégorie",
        "Compétence",
        "Éval 1 Date",
        "Éval 1 Score",
        "Éval 1 Commentaires",
        "Éval 2 Date",
        "Éval 2 Score",
        "Éval 2 Commentaires",
        "Éval 3 Date",
        "Éval 3 Score",
        "Éval 3 Commentaires",
        "Éval 4 Date",
        "Éval 4 Score",
        "Éval 4 Commentaires",
    ]
    builder.add_headers(headers, row=3)

    # Skills
    row_idx = 4
    for category in categories:
        for skill in category.skills:
            ws.cell(row=row_idx, column=1, value=category.name)
            ws.cell(row=row_idx, column=2, value=skill.name)

            fill = get_category_fill(category.id, category_colors)
            if fill:
                ws.cell(row=row_idx, column=1).fill = fill

            row_idx += 1

    # Column widths
    builder.set_column_widths({"A": 25, "B": 35})
    builder.set_column_widths_range(3, 14, 15)

    ws.freeze_panes = "C4"


def generate_xlsx(config: Config, categories: list[Category], output_path: Path):
    """Generate the complete Excel skills matrix."""
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Create all sheets
    create_levels_sheet(wb, config)
    create_matrix_sheet(wb, config, categories)
    create_level_descriptions_sheet(wb, config, categories)
    create_assessment_sheet(wb, config, categories)
    create_radar_data_sheet(wb, config, categories)
    create_development_plan_sheet(wb, config)
    create_history_sheet(wb, config, categories)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save
    wb.save(output_path)

    return {
        "total_skills": sum(len(cat.skills) for cat in categories),
        "total_categories": len(categories),
        "total_roles": len(config.roles),
        "total_career_levels": len(config.career_levels),
    }
