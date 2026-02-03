"""
Integration tests for Excel generator.
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import load_workbook

from skills_matrix.loader import load_config, load_all_skills
from skills_matrix.generator import generate_xlsx


class TestGenerateXlsx:
    """Integration tests for the Excel generation."""

    @pytest.fixture
    def generated_workbook(self):
        """Generate a workbook using actual data files."""
        data_path = Path("data")
        config = load_config(data_path / "config.yaml")
        categories = load_all_skills(data_path / "skills", config.category_order)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = Path(f.name)

        stats = generate_xlsx(config, categories, output_path)
        wb = load_workbook(output_path)

        yield wb, stats, config, categories

        # Cleanup
        output_path.unlink(missing_ok=True)

    def test_workbook_has_7_sheets(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        assert len(wb.sheetnames) == 7

    def test_sheet_names_are_correct(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        expected_sheets = [
            "Niveaux",
            "Matrice Compétences",
            "Descriptions par Niveau",
            "Mon Évaluation",
            "Profils par Rôle (Radar)",
            "Plan Développement",
            "Historique Évaluations",
        ]
        assert wb.sheetnames == expected_sheets

    def test_stats_are_correct(self, generated_workbook):
        wb, stats, config, categories = generated_workbook

        total_skills = sum(len(cat.skills) for cat in categories)
        assert stats["total_skills"] == total_skills
        assert stats["total_categories"] == len(categories)
        assert stats["total_roles"] == len(config.roles)
        assert stats["total_career_levels"] == len(config.career_levels)

    def test_levels_sheet_has_content(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Niveaux"]

        # Check title
        assert ws["A1"].value == "Échelle de Compétences"

        # Check headers exist
        assert ws["A3"].value == "Niveau"
        assert ws["B3"].value == "Description"

    def test_matrix_sheet_has_skills(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Matrice Compétences"]

        # Check headers
        assert ws["A1"].value == "Catégorie"
        assert ws["B1"].value == "Compétence"
        assert ws["C1"].value == "Description"

        # Check first skill is present (row 2)
        first_category = categories[0]
        first_skill = first_category.skills[0]
        assert ws["A2"].value == first_category.name
        assert ws["B2"].value == first_skill.name

    def test_matrix_sheet_has_all_roles(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Matrice Compétences"]

        # Check that role headers exist
        header_row = [
            ws.cell(row=1, column=col).value
            for col in range(1, 50)
            if ws.cell(row=1, column=col).value
        ]

        for role in config.roles:
            # Each role should have C/S and 4 level columns
            assert any(role.name in str(h) for h in header_row), (
                f"Role {role.name} not found in headers"
            )

    def test_assessment_sheet_has_dropdowns(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Mon Évaluation"]

        # Check title
        assert ws["A1"].value == "Auto-évaluation des Compétences"

        # Check data validations exist
        assert len(ws.data_validations.dataValidation) > 0

    def test_development_plan_has_structure(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Plan Développement"]

        # Check title
        assert ws["A1"].value == "Plan de Développement Annuel"

        # Check headers
        assert ws["A4"].value == "Compétence à Développer"
        assert ws["H4"].value == "Statut"

    def test_history_sheet_has_all_skills(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Historique Évaluations"]

        # Count skill rows (starting from row 4)
        total_skills = sum(len(cat.skills) for cat in categories)
        skill_rows = 0
        for row in range(4, 4 + total_skills + 10):
            if ws.cell(row=row, column=2).value:  # Check skill name column
                skill_rows += 1

        assert skill_rows == total_skills

    def test_descriptions_sheet_has_level_columns(self, generated_workbook):
        wb, stats, config, categories = generated_workbook
        ws = wb["Descriptions par Niveau"]

        # Check level headers
        assert ws["C1"].value == "Niveau 0"
        assert ws["D1"].value == "Niveau 1"
        assert ws["E1"].value == "Niveau 2"
        assert ws["F1"].value == "Niveau 3"
        assert ws["G1"].value == "Niveau 4"


class TestGenerateXlsxWithFixtures:
    """Tests using pytest fixtures for isolated testing."""

    def test_generate_with_minimal_data(
        self, sample_config_data, sample_skill_data, tmp_path
    ):
        """Test generation with minimal fixture data."""
        from skills_matrix.models import (
            Config,
            Role,
            CareerLevel,
            SkillLevel,
            ColorConfig,
            Category,
            Skill,
        )

        # Build config from fixture
        roles = [Role(**r) for r in sample_config_data["roles"]]
        career_levels = [
            CareerLevel(**cl) for cl in sample_config_data["career_levels"]
        ]
        skill_levels = {
            k: [SkillLevel(**sl) for sl in v]
            for k, v in sample_config_data["skill_levels"].items()
        }
        colors = ColorConfig(**sample_config_data["colors"])
        config = Config(
            roles=roles,
            career_levels=career_levels,
            skill_levels=skill_levels,
            colors=colors,
            output=sample_config_data["output"],
            category_order=sample_config_data["category_order"],
        )

        # Build category from fixture
        skill_data = sample_skill_data["skills"][0]
        level_descriptions = {
            int(k): v for k, v in skill_data["level_descriptions"].items()
        }
        skill = Skill(
            id=skill_data["id"],
            name=skill_data["name"],
            description=skill_data["description"],
            core_roles=skill_data["core_roles"],
            levels=skill_data["levels"],
            level_descriptions=level_descriptions,
            resources=skill_data.get("resources"),
        )
        category = Category(
            id=sample_skill_data["category"]["id"],
            name=sample_skill_data["category"]["name"],
            skills=[skill],
        )

        # Generate
        output_path = tmp_path / "test_output.xlsx"
        stats = generate_xlsx(config, [category], output_path)

        # Verify
        assert output_path.exists()
        assert stats["total_skills"] == 1
        assert stats["total_categories"] == 1
        assert stats["total_roles"] == 2

        # Load and verify structure
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 7

    def test_output_directory_created(
        self, sample_config_data, sample_skill_data, tmp_path
    ):
        """Test that output directory is created if it doesn't exist."""
        from skills_matrix.models import (
            Config,
            Role,
            CareerLevel,
            SkillLevel,
            ColorConfig,
            Category,
        )

        # Build minimal config and category
        roles = [Role(**r) for r in sample_config_data["roles"]]
        career_levels = [
            CareerLevel(**cl) for cl in sample_config_data["career_levels"]
        ]
        skill_levels = {
            k: [SkillLevel(**sl) for sl in v]
            for k, v in sample_config_data["skill_levels"].items()
        }
        colors = ColorConfig(**sample_config_data["colors"])
        config = Config(
            roles=roles,
            career_levels=career_levels,
            skill_levels=skill_levels,
            colors=colors,
            output={},
        )

        category = Category(id="test", name="Test", skills=[])

        # Output to non-existent directory
        output_path = tmp_path / "new_dir" / "subdir" / "output.xlsx"
        assert not output_path.parent.exists()

        generate_xlsx(config, [category], output_path)

        assert output_path.exists()
        assert output_path.parent.exists()
