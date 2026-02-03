"""
Tests for Excel styling utilities.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border

from skills_matrix.styles import (
    create_fill,
    create_thin_border,
    create_header_style,
    apply_header_style,
    get_level_fill,
    get_category_fill,
)


class TestCreateFill:
    def test_create_fill_returns_pattern_fill(self):
        fill = create_fill("FF0000")
        assert isinstance(fill, PatternFill)
        assert fill.start_color.rgb == "00FF0000"
        assert fill.fill_type == "solid"

    def test_create_fill_cached(self):
        # Same color should return cached instance
        fill1 = create_fill("00FF00")
        fill2 = create_fill("00FF00")
        assert fill1 is fill2

    def test_create_fill_different_colors(self):
        fill1 = create_fill("FF0000")
        fill2 = create_fill("00FF00")
        assert fill1 is not fill2


class TestCreateThinBorder:
    def test_create_thin_border_returns_border(self):
        border = create_thin_border()
        assert isinstance(border, Border)
        assert border.left.style == "thin"
        assert border.right.style == "thin"
        assert border.top.style == "thin"
        assert border.bottom.style == "thin"

    def test_border_color(self):
        border = create_thin_border()
        assert border.left.color.rgb == "00CCCCCC"


class TestCreateHeaderStyle:
    def test_header_style_contains_required_keys(self):
        style = create_header_style("3943B4")
        assert "font" in style
        assert "fill" in style
        assert "alignment" in style
        assert "border" in style

    def test_header_font_is_bold_white(self):
        style = create_header_style("3943B4")
        assert style["font"].bold is True
        assert style["font"].color.rgb == "00FFFFFF"

    def test_header_fill_uses_provided_color(self):
        style = create_header_style("FF0000")
        assert style["fill"].start_color.rgb == "00FF0000"


class TestApplyHeaderStyle:
    def test_apply_header_style_to_cell(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = "Header"

        apply_header_style(cell, "3943B4")

        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        assert cell.fill.start_color.rgb == "003943B4"
        assert cell.alignment.horizontal == "center"
        assert cell.border.left.style == "thin"


class TestGetLevelFill:
    def test_get_level_fill_existing_level(self):
        level_colors = {"0": "FFCFC9", "1": "FFE5A0", "2": "D4EDBC"}
        fill = get_level_fill("1", level_colors)
        assert fill is not None
        assert fill.start_color.rgb == "00FFE5A0"

    def test_get_level_fill_nonexistent_level(self):
        level_colors = {"0": "FFCFC9", "1": "FFE5A0"}
        fill = get_level_fill("99", level_colors)
        assert fill is None

    def test_get_level_fill_with_int(self):
        level_colors = {"0": "FFCFC9", "1": "FFE5A0"}
        fill = get_level_fill(1, level_colors)  # int instead of str
        assert fill is not None
        assert fill.start_color.rgb == "00FFE5A0"

    def test_get_level_fill_nc(self):
        level_colors = {"NC": "EFEFEF"}
        fill = get_level_fill("NC", level_colors)
        assert fill is not None
        assert fill.start_color.rgb == "00EFEFEF"


class TestGetCategoryFill:
    def test_get_category_fill_existing(self):
        category_colors = {"data_analysis": "FFC8AA", "ml": "E6CFF2"}
        fill = get_category_fill("data_analysis", category_colors)
        assert fill is not None
        assert fill.start_color.rgb == "00FFC8AA"

    def test_get_category_fill_nonexistent(self):
        category_colors = {"data_analysis": "FFC8AA"}
        fill = get_category_fill("nonexistent", category_colors)
        assert fill is None
